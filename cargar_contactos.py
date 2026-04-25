import openpyxl, requests

URL = "https://vazneyvdycwfhqipyxdi.supabase.co"
KEY = "sb_secret_X_1CDqRGfwaKXB2QIJAPeA_BFfxpPYa"
HEADERS = {
    "apikey": KEY,
    "Authorization": f"Bearer {KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal"
}

BASE = r"C:\Users\nn\Desktop\base de datos"

def clean_tel(t):
    if not t: return None
    t = str(t).replace('+','').replace(' ','').replace('-','').strip()
    if t.startswith('56') and len(t) == 11: return t
    if t.startswith('9') and len(t) == 9: return '56' + t
    return None

def title(s):
    if not s: return ''
    return str(s).title().strip()

contacts = {}

print("Leyendo DirecTV...")
wb = openpyxl.load_workbook(BASE + r'\base_datos_priorizada.xlsx', read_only=True)
ws = wb.active
for row in ws.iter_rows(min_row=2, values_only=True):
    tel = clean_tel(row[5])
    if not tel: continue
    contacts[tel] = {
        'telefono': tel, 'telefono_2': clean_tel(row[6]),
        'nombre': title(row[4]), 'comuna': title(row[7]),
        'region': str(row[8] or ''), 'empresa_actual': 'DirecTV',
        'producto_actual': str(row[12] or ''), 'fuente': 'DirecTV Ventas Caidas',
        'email': str(row[13] or ''), 'prioridad': int(row[0]) if row[0] else 3,
        'pipeline_stage': 'no_contactado', 'heat_score': 0, 'temperatura': 'frio'
    }
wb.close()
print(f"DirecTV: {len(contacts)}")

print("Leyendo Movistar...")
wb = openpyxl.load_workbook(BASE + r'\luis barrios movistar.xlsx', read_only=True)
ws = wb.active
for row in ws.iter_rows(min_row=2, values_only=True):
    tel = clean_tel(row[1])
    if not tel or tel in contacts: continue
    contacts[tel] = {
        'telefono': tel, 'nombre': title(row[7]),
        'comuna': title(row[3]), 'region': '',
        'empresa_actual': 'Movistar', 'producto_actual': str(row[8] or ''),
        'fuente': 'Movistar Base', 'prioridad': 2,
        'pipeline_stage': 'no_contactado', 'heat_score': 0, 'temperatura': 'frio',
        'direccion': str(row[6] or '')
    }
wb.close()
print(f"Total tras Movistar: {len(contacts)}")

print("Leyendo EMATOS...")
wb = openpyxl.load_workbook(BASE + r'\TOTA Fija EMATOS 080525.xlsx', read_only=True)
ws = wb.active
for row in ws.iter_rows(min_row=2, values_only=True):
    tel = clean_tel(row[6])
    if not tel: continue
    if tel in contacts:
        if row[8]: contacts[tel]['comuna'] = title(row[8])
        if row[21]: contacts[tel]['region'] = str(row[21])
        continue
    contacts[tel] = {
        'telefono': tel, 'nombre': title(row[12]),
        'comuna': title(row[8]), 'region': str(row[21] or ''),
        'empresa_actual': 'Movistar', 'producto_actual': str(row[5] or ''),
        'fuente': 'Movistar EMATOS', 'prioridad': 2,
        'pipeline_stage': 'no_contactado', 'heat_score': 0, 'temperatura': 'frio',
        'direccion': str(row[11] or '')
    }
wb.close()
print(f"Total final: {len(contacts)}")

data = list(contacts.values())
total = 0
for i in range(0, len(data), 500):
    batch = data[i:i+500]
    r = requests.post(f"{URL}/rest/v1/contacts", headers=HEADERS, json=batch)
    if r.status_code in (200, 201):
        total += len(batch)
        print(f"Subidos: {total}/{len(data)}")
    else:
        print(f"ERROR: {r.status_code} - {r.text[:200]}")

print(f"COMPLETADO: {total} contactos")
