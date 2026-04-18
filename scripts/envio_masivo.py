# scripts/envio_masivo.py — Envío masivo de plantilla WhatsApp desde Excel
# Conexion Sin Limites

"""
Lee un archivo Excel con columnas 'cliente', 'tel_limpio' y 'prioridad',
y envía la plantilla 'bienvenida_conexion' a cada número via Meta Cloud API.

Antes de enviar consulta la tabla `envios_realizados` en PostgreSQL y descarta
los números a los que ya se envió esta plantilla. Después de cada sesión
registra los envíos exitosos en esa tabla para que la próxima sesión los omita.

Uso:
    python scripts/envio_masivo.py archivo.xlsx [--prioridad N] [--limite N] [--db-url URL]

    --prioridad  Filtra solo filas con ese valor en columna 'prioridad' (default: 1)
    --limite     Máximo de contactos NUEVOS a enviar (default: 100)
    --db-url     URL PostgreSQL de produccion (sobreescribe PROD_DATABASE_URL del .env)

El Excel debe tener estas columnas (primera fila = encabezados):
    cliente           | tel_limpio    | prioridad
    Juan Pérez        | 56912345678   | 1
    María García      | 56987654321   | 2

IMPORTANTE:
- tel_limpio debe incluir código de país SIN el signo +  (ej: 56978016298)
- La plantilla 'bienvenida_conexion' debe estar aprobada en Meta
- El parámetro {{1}} usa el primer nombre + primer apellido del contacto
"""

import os
import sys
import time
import asyncio
import argparse
import httpx
import asyncpg
import openpyxl
from datetime import datetime
from dotenv import load_dotenv

# Ruta absoluta al .env del proyecto, funciona desde cualquier directorio.
# override=True evita que variables del sistema sobreescriban las del .env.
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(_PROJECT_ROOT, ".env"), override=True)

# ── Configuración ──────────────────────────────────────────────
ACCESS_TOKEN    = "".join((os.getenv("META_ACCESS_TOKEN") or "").split())
PHONE_NUMBER_ID = "".join((os.getenv("META_PHONE_NUMBER_ID") or "").split())
TEMPLATE_NAME   = "bienvenida_conexion"
TEMPLATE_LANG   = "es_CL"
API_VERSION     = "v21.0"
PAUSA_SEGUNDOS  = 1
# ──────────────────────────────────────────────────────────────


# ── URL de base de datos ───────────────────────────────────────

def _resolver_db_url(db_url_arg: str = "") -> str:
    """
    Resuelve la URL de BD con esta prioridad:
      1. --db-url (argumento CLI)
      2. PROD_DATABASE_URL (.env) — BD de produccion Railway
      3. DATABASE_URL (.env)      — fallback desarrollo
    """
    from urllib.parse import urlparse

    raw = (
        db_url_arg.strip()
        or os.getenv("PROD_DATABASE_URL", "").strip()
        or os.getenv("DATABASE_URL", "").strip()
    )
    if not raw:
        return ""

    url = raw.replace("postgresql+asyncpg://", "postgresql://")

    try:
        fuente = (
            "--db-url"          if db_url_arg.strip()
            else "PROD_DATABASE_URL" if os.getenv("PROD_DATABASE_URL", "").strip()
            else "DATABASE_URL"
        )
        host = urlparse(url).hostname or url[:30]
        print(f"BD : {host}  [{fuente}]")
    except Exception:
        pass

    return url


# ── Tabla envios_realizados ────────────────────────────────────

async def preparar_tabla(db_url: str):
    """
    Crea la tabla envios_realizados si no existe.
    La restricción UNIQUE (telefono, plantilla) garantiza que no se
    registre el mismo número dos veces para la misma plantilla.
    """
    conn = await asyncpg.connect(db_url)
    await conn.execute("""
        CREATE TABLE IF NOT EXISTS envios_realizados (
            id          SERIAL PRIMARY KEY,
            telefono    TEXT        NOT NULL,
            plantilla   TEXT        NOT NULL,
            exitoso     BOOLEAN     NOT NULL DEFAULT TRUE,
            enviado_en  TIMESTAMP   NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """)
    await conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS uq_envios_tel_plantilla
        ON envios_realizados (telefono, plantilla)
    """)
    await conn.close()


async def obtener_telefonos_ya_enviados(db_url: str, plantilla: str) -> set[str]:
    """
    Devuelve el set de teléfonos a los que ya se envió esta plantilla
    (exitoso = TRUE). Los fallidos no se incluyen — pueden reintentarse.
    Normaliza los valores de la BD para que el match sea consistente.
    """
    conn = await asyncpg.connect(db_url)
    rows = await conn.fetch(
        "SELECT telefono FROM envios_realizados WHERE plantilla = $1 AND exitoso = TRUE",
        plantilla,
    )
    await conn.close()
    # Normalizar por las dudas (quitar +, espacios, guiones residuales en BD)
    return {r["telefono"].replace("+", "").replace(" ", "").replace("-", "").strip()
            for r in rows}


async def registrar_envios(db_url: str, resultados: list[dict], plantilla: str):
    """
    Inserta en envios_realizados los envíos de esta sesión.
    ON CONFLICT DO NOTHING — si un teléfono ya estaba registrado se ignora.
    Se registran exitosos Y fallidos para auditoría completa.
    """
    registros = [
        (r["telefono"], plantilla, r["estado"] == "enviado")
        for r in resultados
        if r.get("estado") in ("enviado", "error")   # excluye "saltado"
    ]
    if not registros:
        return

    conn = await asyncpg.connect(db_url)
    await conn.executemany(
        """
        INSERT INTO envios_realizados (telefono, plantilla, exitoso)
        VALUES ($1, $2, $3)
        ON CONFLICT (telefono, plantilla) DO NOTHING
        """,
        registros,
    )
    await conn.close()
    print(f"Registrados en BD: {len(registros)} envíos (exitosos + fallidos).")


# ── Lectura del Excel ──────────────────────────────────────────

def primer_nombre_apellido(nombre_completo: str) -> str:
    """
    Extrae primer nombre + primer apellido de un nombre completo.
      1 palabra  → "Juan"
      2 palabras → "Juan Pérez"
      3 palabras → "Juan Pérez González"  → "Juan Pérez"
      4+palabras → "Juan Carlos Pérez González" → "Juan Pérez"
    """
    partes = nombre_completo.strip().split()
    if len(partes) <= 2:
        return " ".join(partes)
    elif len(partes) == 3:
        return f"{partes[0]} {partes[1]}"
    else:
        return f"{partes[0]} {partes[2]}"


def normalizar_telefono(raw) -> str:
    """
    Convierte cualquier valor de celda a string de teléfono limpio.
    Maneja números flotantes de Excel (ej: 56912345678.0 -> '56912345678').
    """
    if raw is None:
        return ""
    # Excel puede guardar números como float (56912345678.0)
    if isinstance(raw, float):
        raw = str(int(raw))
    else:
        raw = str(raw)
    return raw.replace("+", "").replace(" ", "").replace("-", "").strip()


def leer_excel(ruta: str, prioridad: int = 1, desde: int = 1) -> list[dict]:
    """
    Lee todos los contactos con la prioridad indicada.
    'desde' es el número de orden (1-based) dentro de los contactos
    que pasan el filtro de prioridad. desde=1 = todos, desde=201 = saltar primeros 200.
    El límite se aplica en main() DESPUES de filtrar los ya enviados.
    """
    wb = openpyxl.load_workbook(ruta)
    ws = wb.active

    encabezados = {
        str(cell.value).strip().lower(): idx
        for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)
    }

    requeridas = ["cliente", "tel_limpio", "prioridad"]
    faltantes  = [c for c in requeridas if c not in encabezados]
    if faltantes:
        print(f"ERROR: Faltan columnas en el Excel: {', '.join(faltantes)}")
        sys.exit(1)

    col_nombre    = encabezados["cliente"]
    col_telefono  = encabezados["tel_limpio"]
    col_prioridad = encabezados["prioridad"]

    contactos = []
    omitidos  = 0
    for fila in ws.iter_rows(min_row=2, values_only=True):
        nombre     = str(fila[col_nombre - 1]    or "").strip()
        prio_valor = fila[col_prioridad - 1]

        try:
            if int(prio_valor) != prioridad:
                omitidos += 1
                continue
        except (TypeError, ValueError):
            omitidos += 1
            continue

        # Normalizar teléfono (maneja floats de Excel como 56912345678.0)
        telefono = normalizar_telefono(fila[col_telefono - 1])

        if nombre and telefono:
            contactos.append({"nombre": nombre, "telefono": telefono})

    print(f"Filas omitidas (prioridad != {prioridad}): {omitidos}")

    # Aplicar --desde: saltar los primeros (desde-1) contactos
    if desde > 1:
        saltados_desde = desde - 1
        contactos = contactos[saltados_desde:]
        print(f"Contactos saltados por --desde {desde}       : {saltados_desde}")

    return contactos


# ── Envío ──────────────────────────────────────────────────────

def enviar_plantilla(cliente: httpx.Client, nombre: str, telefono: str) -> tuple[bool, str]:
    """Envía la plantilla a un número. Retorna (ok, detalle)."""
    url = f"https://graph.facebook.com/{API_VERSION}/{PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json",
    }
    payload = {
        "messaging_product": "whatsapp",
        "to": telefono,
        "type": "template",
        "template": {
            "name": TEMPLATE_NAME,
            "language": {"code": TEMPLATE_LANG},
            "components": [{"type": "body", "parameters": [{"type": "text", "text": nombre}]}],
        },
    }
    try:
        r = cliente.post(url, json=payload, headers=headers, timeout=15)
        if r.status_code == 200:
            msg_id = r.json().get("messages", [{}])[0].get("id", "ok")
            return True, msg_id
        else:
            error = r.json().get("error", {}).get("message", r.text)
            return False, error
    except Exception as e:
        return False, str(e)


# ── Log Excel ──────────────────────────────────────────────────

def guardar_log(resultados: list[dict]) -> str:
    """Guarda un Excel con los resultados completos de la sesión."""
    nombre_log = f"envio_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.append(["nombre_completo", "nombre_plantilla", "telefono",
               "estado", "detalle", "timestamp"])
    for r in resultados:
        ws.append([r["nombre"], r["nombre_plantilla"], r["telefono"],
                   r["estado"], r["detalle"], r["timestamp"]])
    wb.save(nombre_log)
    return nombre_log


# ── Main ───────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Envio masivo de plantilla WhatsApp",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
BD de produccion (prioridad):
  1. --db-url          URL directa en el comando
  2. PROD_DATABASE_URL Variable en .env  <-- Railway prod
  3. DATABASE_URL      Variable en .env  (fallback)
        """,
    )
    parser.add_argument("archivo",       help="Ruta al archivo Excel")
    parser.add_argument("--prioridad",   type=int, default=1,
                        help="Prioridad a filtrar (default: 1)")
    parser.add_argument("--desde",       type=int, default=1,
                        help="Fila de inicio dentro de los contactos con esa prioridad "
                             "(1-based, default: 1 = todos). Ej: --desde 201 salta los primeros 200.")
    parser.add_argument("--limite",      type=int, default=100,
                        help="Maximo de contactos NUEVOS a enviar (default: 100)")
    parser.add_argument("--db-url",      default="",
                        help="URL PostgreSQL de produccion (sobreescribe .env)")
    args = parser.parse_args()

    if not os.path.exists(args.archivo):
        print(f"ERROR: No se encontro el archivo '{args.archivo}'")
        sys.exit(1)

    if not ACCESS_TOKEN or not PHONE_NUMBER_ID:
        print("ERROR: META_ACCESS_TOKEN o META_PHONE_NUMBER_ID no configurados en .env")
        sys.exit(1)

    # ── 1. Resolver BD ─────────────────────────────────────────
    db_url = _resolver_db_url(args.db_url)
    if not db_url:
        print("ADVERTENCIA: Sin DB configurada — se enviara a todos sin filtrar duplicados.")

    # ── 2. Crear tabla si no existe ────────────────────────────
    if db_url:
        asyncio.run(preparar_tabla(db_url))

    # ── 3. Leer Excel ──────────────────────────────────────────
    todos = leer_excel(args.archivo, prioridad=args.prioridad, desde=args.desde)
    print(f"\nContactos en Excel (prioridad={args.prioridad}, desde={args.desde}): {len(todos)}")

    # ── 4. Filtrar ya enviados ─────────────────────────────────
    if db_url:
        print(f"Consultando envios_realizados para plantilla '{TEMPLATE_NAME}'...")
        ya_enviados = asyncio.run(obtener_telefonos_ya_enviados(db_url, TEMPLATE_NAME))
        print(f"Ya enviados anteriormente   : {len(ya_enviados)}")
    else:
        ya_enviados = set()

    saltados = [c for c in todos if c["telefono"] in ya_enviados]
    nuevos   = [c for c in todos if c["telefono"] not in ya_enviados]

    print(f"Saltados (ya enviados)       : {len(saltados)}")
    print(f"Nuevos para esta sesion      : {len(nuevos)}")

    # ── 5. Aplicar límite sobre nuevos ─────────────────────────
    contactos = nuevos[:args.limite]
    total     = len(contactos)

    print(f"\nContactos a enviar (limite={args.limite}): {total}")
    print(f"Plantilla : {TEMPLATE_NAME} | Idioma: {TEMPLATE_LANG}")
    print(f"Phone ID  : {PHONE_NUMBER_ID}")
    print("-" * 52)

    if total == 0:
        print("No hay contactos nuevos para enviar.")
        sys.exit(0)

    confirmacion = input(f"Enviar a {total} contactos? (si/no): ").strip().lower()
    if confirmacion != "si":
        print("Cancelado.")
        sys.exit(0)

    # ── 6. Enviar ──────────────────────────────────────────────
    print()
    resultados = []
    exitosos   = 0
    fallidos   = 0

    with httpx.Client() as cliente:
        for i, contacto in enumerate(contactos, start=1):
            nombre           = contacto["nombre"]
            telefono         = contacto["telefono"]
            nombre_plantilla = primer_nombre_apellido(nombre)

            ok, detalle = enviar_plantilla(cliente, nombre_plantilla, telefono)
            estado      = "enviado" if ok else "error"
            ts          = datetime.now().strftime("%H:%M:%S")

            resultados.append({
                "nombre":           nombre,
                "telefono":         telefono,
                "nombre_plantilla": nombre_plantilla,
                "estado":           estado,
                "detalle":          detalle,
                "timestamp":        ts,
            })

            print(f"[{i}/{total}] {'OK   ' if ok else 'ERROR'} "
                  f"{nombre} -> '{nombre_plantilla}' ({telefono}) {detalle}")

            if ok:
                exitosos += 1
            else:
                fallidos += 1

            if i < total:
                time.sleep(PAUSA_SEGUNDOS)

    # ── 7. Registrar en BD ─────────────────────────────────────
    if db_url:
        asyncio.run(registrar_envios(db_url, resultados, TEMPLATE_NAME))

    # ── 8. Agregar saltados al log y guardar ───────────────────
    ts_fin = datetime.now().strftime("%H:%M:%S")
    for c in saltados:
        resultados.append({
            "nombre":           c["nombre"],
            "telefono":         c["telefono"],
            "nombre_plantilla": primer_nombre_apellido(c["nombre"]),
            "estado":           "saltado",
            "detalle":          f"ya enviado ({TEMPLATE_NAME})",
            "timestamp":        ts_fin,
        })

    archivo_log = guardar_log(resultados)

    print()
    print("=" * 52)
    print(f"  Envio completado")
    print(f"  Exitosos : {exitosos}")
    print(f"  Fallidos : {fallidos}")
    print(f"  Saltados : {len(saltados)}  (ya enviados en sesiones anteriores)")
    print(f"  Log      : {archivo_log}")
    print("=" * 52)


if __name__ == "__main__":
    main()
