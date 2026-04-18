# scripts/migrar_leads_enviados.py
# Conexion Sin Limites
#
# Lee los 295 teléfonos de envios_realizados, los cruza con el Excel
# y hace upsert en la tabla leads con los datos del contacto.
#
# Uso:
#   python scripts/migrar_leads_enviados.py <ruta_excel> [--dry-run]
#
#   --dry-run  Muestra lo que haría sin escribir nada en la BD.

import os
import sys
import asyncio
import argparse
import openpyxl
import asyncpg
from dotenv import load_dotenv

sys.stdout.reconfigure(encoding="utf-8")

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(_PROJECT_ROOT, ".env"), override=True)

PLANTILLA       = "bienvenida_conexion"
CLIENTE_ID      = 1
ESTADO_NUEVO    = "contactado"
ORIGEN          = "envio_masivo"
AGENTE          = "valentina"

# Estados que NO deben bajarse a 'contactado' en el ON CONFLICT
ESTADOS_AVANZADOS = ("interesado", "tibio", "caliente", "direccion_obtenida",
                     "listo_para_cierre", "cerrado", "seguimiento", "no_interesado")


def normalizar_tel(raw) -> str:
    if raw is None:
        return ""
    if isinstance(raw, float):
        raw = str(int(raw))
    else:
        raw = str(raw)
    return raw.replace("+", "").replace(" ", "").replace("-", "").strip()


def leer_excel(ruta: str) -> dict[str, dict]:
    """
    Devuelve un dict {telefono_normalizado: {nombre, ciudad, region, prioridad}}
    con TODOS los contactos del Excel (todas las prioridades).
    """
    wb = openpyxl.load_workbook(ruta)
    ws = wb.active

    headers = {
        str(cell.value).strip().lower(): idx
        for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)
    }

    requeridas = ["cliente", "tel_limpio"]
    faltantes  = [c for c in requeridas if c not in headers]
    if faltantes:
        print(f"ERROR: Faltan columnas en el Excel: {', '.join(faltantes)}")
        sys.exit(1)

    col_nombre   = headers["cliente"]
    col_tel      = headers["tel_limpio"]
    col_prio     = headers.get("prioridad")
    col_ciudad   = headers.get("ciudad")
    col_region   = headers.get("region")

    contactos = {}
    for fila in ws.iter_rows(min_row=2, values_only=True):
        nombre   = str(fila[col_nombre - 1] or "").strip()
        telefono = normalizar_tel(fila[col_tel - 1])
        if not nombre or not telefono:
            continue

        prio = 1
        if col_prio:
            try:
                prio = int(fila[col_prio - 1])
            except (TypeError, ValueError):
                prio = 1

        ciudad = str(fila[col_ciudad - 1] or "").strip() if col_ciudad else ""
        region = str(fila[col_region - 1] or "").strip() if col_region else ""

        contactos[telefono] = {
            "nombre":    nombre,
            "ciudad":    ciudad,
            "region":    region,
            "prioridad": prio,
        }

    return contactos


async def main(excel_path: str, dry_run: bool):
    db_url = (
        os.getenv("PROD_DATABASE_URL", "").strip()
        or os.getenv("DATABASE_URL", "").strip()
    )
    if not db_url:
        print("ERROR: No se encontro PROD_DATABASE_URL ni DATABASE_URL en .env")
        sys.exit(1)

    db_url = db_url.replace("postgresql+asyncpg://", "postgresql://")

    print(f"BD     : {db_url.split('@')[-1]}")
    print(f"Excel  : {excel_path}")
    print(f"Modo   : {'DRY RUN (sin escritura)' if dry_run else 'ESCRITURA REAL'}")
    print("-" * 60)

    # 1. Leer Excel
    contactos_excel = leer_excel(excel_path)
    print(f"Contactos en Excel        : {len(contactos_excel)}")

    # 2. Leer teléfonos de envios_realizados
    conn = await asyncpg.connect(db_url)
    rows = await conn.fetch(
        "SELECT telefono FROM envios_realizados WHERE plantilla = $1 AND exitoso = TRUE",
        PLANTILLA,
    )
    enviados = {
        r["telefono"].replace("+", "").replace(" ", "").replace("-", "").strip()
        for r in rows
    }
    print(f"Telefonos en envios_realizados : {len(enviados)}")

    # 3. Cruzar
    a_insertar = []
    sin_datos  = []
    for tel in sorted(enviados):
        datos = contactos_excel.get(tel)
        if datos:
            a_insertar.append((tel, datos))
        else:
            sin_datos.append(tel)

    print(f"Con datos en Excel        : {len(a_insertar)}")
    print(f"Sin datos en Excel        : {len(sin_datos)}  (se insertan solo con telefono)")
    print()

    if not a_insertar and not sin_datos:
        print("Nada que migrar.")
        await conn.close()
        return

    if dry_run:
        print("--- PREVIEW (primeros 10) ---")
        for tel, datos in a_insertar[:10]:
            tag = f'["prioridad_{datos["prioridad"]}"]'
            print(f"  UPSERT | {datos['nombre']:<30} | {tel} | {datos['ciudad']} | {datos['region']} | tags={tag}")
        if sin_datos:
            print(f"  ... y {len(sin_datos)} telefonos sin datos en Excel.")
        await conn.close()
        return

    # 4. Upsert en leads
    insertados   = 0
    actualizados = 0
    sin_match    = 0

    # Registros con datos del Excel
    for tel, datos in a_insertar:
        notas = f"Region: {datos['region']}" if datos.get("region") else ""
        tag   = f'["prioridad_{datos["prioridad"]}"]'

        result = await conn.fetchrow(
            """
            INSERT INTO leads (
                telefono, nombre, comuna, notas,
                estado, origen, agente, tags,
                cliente_id, ultima_interaccion, created_at
            )
            VALUES ($1, $2, $3, $4,
                    $5, $6, $7, $8::text,
                    $9, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            ON CONFLICT (telefono) DO UPDATE SET
                nombre  = CASE WHEN leads.nombre IS NULL OR leads.nombre = ''
                               THEN EXCLUDED.nombre  ELSE leads.nombre  END,
                comuna  = CASE WHEN leads.comuna IS NULL OR leads.comuna = ''
                               THEN EXCLUDED.comuna  ELSE leads.comuna  END,
                notas   = CASE WHEN leads.notas  IS NULL OR leads.notas  = ''
                               THEN EXCLUDED.notas   ELSE leads.notas   END,
                tags    = CASE WHEN leads.tags = '[]' OR leads.tags IS NULL
                               THEN EXCLUDED.tags    ELSE leads.tags    END,
                estado  = CASE WHEN lower(leads.estado) = ANY($10::text[])
                               THEN leads.estado     ELSE EXCLUDED.estado END,
                ultima_interaccion = CURRENT_TIMESTAMP
            RETURNING (xmax = 0) AS inserted
            """,
            tel,
            datos["nombre"],
            datos["ciudad"],
            notas,
            ESTADO_NUEVO,
            ORIGEN,
            AGENTE,
            tag,
            CLIENTE_ID,
            list(ESTADOS_AVANZADOS),
        )
        if result and result["inserted"]:
            insertados += 1
        else:
            actualizados += 1

    # Registros sin datos en Excel (solo teléfono)
    for tel in sin_datos:
        result = await conn.fetchrow(
            """
            INSERT INTO leads (
                telefono, nombre,
                estado, origen, agente, tags,
                cliente_id, ultima_interaccion, created_at
            )
            VALUES ($1, $2,
                    $3, $4, $5, '["prioridad_1"]',
                    $6, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            ON CONFLICT (telefono) DO UPDATE SET
                estado = CASE WHEN lower(leads.estado) = ANY($7::text[])
                              THEN leads.estado ELSE EXCLUDED.estado END,
                ultima_interaccion = CURRENT_TIMESTAMP
            RETURNING (xmax = 0) AS inserted
            """,
            tel,
            "Desconocido",
            ESTADO_NUEVO,
            ORIGEN,
            AGENTE,
            CLIENTE_ID,
            list(ESTADOS_AVANZADOS),
        )
        if result and result["inserted"]:
            insertados += 1
            sin_match += 1
        else:
            actualizados += 1

    await conn.close()

    print("=" * 60)
    print(f"  Insertados (nuevos)  : {insertados}")
    print(f"  Actualizados         : {actualizados}")
    if sin_match:
        print(f"  Sin datos Excel      : {sin_match}  (insertados como 'Desconocido')")
    print("=" * 60)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Migra envios_realizados -> tabla leads con datos del Excel"
    )
    parser.add_argument("excel", help="Ruta al archivo Excel con los contactos")
    parser.add_argument("--dry-run", action="store_true",
                        help="Muestra preview sin escribir en la BD")
    args = parser.parse_args()

    if not os.path.exists(args.excel):
        print(f"ERROR: No se encontro el archivo '{args.excel}'")
        sys.exit(1)

    asyncio.run(main(args.excel, args.dry_run))
